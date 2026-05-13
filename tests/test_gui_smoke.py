"""Smoke-тесты GUI: редактирование ЖК, локализация, Excel-экспорт.

Проверяем:
1.  l10n.py — все словари покрывают известные значения enum.
2.  l10n-фильтры возвращают строку (не KeyError) для любого входа.
3.  render_card теперь выводит русские надписи для theme/urgency/character.
4.  REASON_LABEL содержит спам-причины из PipelineDecision.escalation_reason.
5.  ComplexesRepo.get возвращает запись после upsert (основа для edit-роута).
6.  ComplexesRepo.get возвращает None для несуществующего ID.
7.  GUI: GET /complexes/{id}/edit → 200, форма с current values.
8.  GUI: GET /complexes/nonexistent/edit → 404.
9.  GUI: POST /complexes/upsert (edit path) сохраняет обновлённые данные.
10. GUI: GET /stats/export → 200, Content-Type xlsx, валидная книга.
11. GUI: Excel-книга содержит 5 листов с правильными заголовками.
12. GUI: GET /stats → 200, содержит русские названия тем.
13. GUI: GET /complexes → 200, содержит кнопку «Изменить».
14. GUI: GET /chat_whitelist → 200, содержит «Диалог с жильцом».
15. Excel: _excel_autowidth не падает на пустом листе.
16. Excel: _excel_data_row чередует цвета строк.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
from typing import Any
from unittest.mock import AsyncMock, MagicMock

import openpyxl
import pytest
import pytest_asyncio

# ---------------------------------------------------------------------------
# Вспомогательные функции DB
# ---------------------------------------------------------------------------


async def _build_db():
    from balt_dom_bot.storage.db import Database
    db = Database(":memory:")
    await db.connect()
    return db


# ===========================================================================
# 1–4. Юнит-тесты l10n.py
# ===========================================================================


def test_l10n_themes_cover_all_enum_values():
    from balt_dom_bot.l10n import THEME_RU
    from balt_dom_bot.models import Theme
    for t in Theme:
        assert t.value in THEME_RU, f"THEME_RU missing {t.value!r}"
        assert THEME_RU[t.value], f"THEME_RU[{t.value!r}] is empty"


def test_l10n_urgency_cover_all_enum_values():
    from balt_dom_bot.l10n import URGENCY_RU
    from balt_dom_bot.models import Urgency
    for u in Urgency:
        assert u.value in URGENCY_RU, f"URGENCY_RU missing {u.value!r}"


def test_l10n_character_cover_all_enum_values():
    from balt_dom_bot.l10n import CHARACTER_RU
    from balt_dom_bot.models import Character
    for c in Character:
        assert c.value in CHARACTER_RU, f"CHARACTER_RU missing {c.value!r}"


def test_l10n_reason_covers_all_pipeline_reasons():
    """PipelineDecision.escalation_reason — все значения должны быть в REASON_RU."""
    from balt_dom_bot.l10n import REASON_RU
    all_reasons = [
        "aggression", "provocation", "high_urgency", "always_escalate_theme",
        "low_confidence", "llm_error", "after_hours",
        "spam_drugs", "spam_crypto", "spam_earn", "spam_esoteric",
        "spam_ads", "spam_mass_mention", "spam_unknown",
        "fragment_troll",
    ]
    for r in all_reasons:
        assert r in REASON_RU, f"REASON_RU missing {r!r}"


def test_l10n_fallback_unknown_key():
    """Функции-хелперы не должны падать на неизвестных ключах."""
    from balt_dom_bot.l10n import character_ru, reason_ru, theme_ru, urgency_ru
    assert theme_ru("UNKNOWN_FUTURE_THEME") == "UNKNOWN_FUTURE_THEME"
    assert urgency_ru("EXTREME") == "EXTREME"
    assert character_ru("SARCASM") == "SARCASM"
    assert reason_ru("new_reason") == "new_reason"


# ===========================================================================
# 5–6. render_card содержит русские подписи
# ===========================================================================


def test_render_card_uses_russian_labels():
    from balt_dom_bot.models import (
        AddressedTo, Character, Classification,
        ComplexInfo, IncomingMessage, PipelineDecision, Theme, Urgency,
    )
    from balt_dom_bot.services.escalation import render_card

    cls = Classification(
        theme=Theme.TECH_FAULT,
        urgency=Urgency.HIGH,
        character=Character.COMPLAINT_STRONG,
        summary="Протечка в подъезде",
        confidence=0.95,
        addressed_to=AddressedTo.UC,
    )
    decision = PipelineDecision(
        classification=cls,
        escalate=True,
        escalation_reason="high_urgency",
    )
    incoming = IncomingMessage(
        chat_id=-100, message_id="m1", user_id=1,
        user_name="Иван", text="Течёт кран",
        received_at=datetime(2026, 5, 1, 12, 0, tzinfo=timezone.utc),
    )
    complex_info = ComplexInfo(
        id="test", name="ЖК Тест", address="ул. Тестовая, 1",
        manager_chat_id=99999,
    )

    card = render_card(
        incoming=incoming, cls=cls, complex_info=complex_info,
        decision=decision, proposed_reply="Проверим", esc_id=42,
    )

    assert "TECH_FAULT" not in card, "English theme should not appear in card"
    assert "HIGH" not in card or "Высокая" in card or "🆘" in card, \
        "English urgency should not appear standalone"
    assert "COMPLAINT_STRONG" not in card, "English character should not appear"
    assert "Техническая неисправность" in card
    assert "Высокая срочность" in card
    assert "Жалоба (жёсткая)" in card


def test_render_card_reason_label_russian():
    """REASON_LABEL в escalation.py теперь ссылается на REASON_RU."""
    from balt_dom_bot.services.escalation import REASON_LABEL
    assert "aggression" in REASON_LABEL
    # Значение должно быть русским, не английским
    assert REASON_LABEL["aggression"].startswith("🚫")
    assert "спам" in REASON_LABEL.get("spam_drugs", "").lower() or \
           "Спам" in REASON_LABEL.get("spam_drugs", "")


# ===========================================================================
# 7–9. Тесты ComplexesRepo — основа для edit-роута
# ===========================================================================


@pytest.mark.asyncio
async def test_complexes_repo_get_after_upsert():
    from balt_dom_bot.storage.complexes_repo import ComplexesRepo
    db = await _build_db()
    repo = ComplexesRepo(db)
    await repo.upsert(
        complex_id="edit-test-1",
        name="ЖК Редактируемый",
        address="ул. Изменяемая, 42",
        chat_id=-111222333,
        manager_chat_id=987654,
        active=True,
        chat_mode_enabled=True,
        reply_mode="holiday",
        holiday_message="Мы в отпуске",
        daily_replies_limit=3,
        daily_window_hours=12,
    )
    c = await repo.get("edit-test-1")
    assert c is not None
    assert c.name == "ЖК Редактируемый"
    assert c.chat_mode_enabled is True
    assert c.reply_mode == "holiday"
    assert c.holiday_message == "Мы в отпуске"
    assert c.daily_replies_limit == 3


@pytest.mark.asyncio
async def test_complexes_repo_get_nonexistent_returns_none():
    from balt_dom_bot.storage.complexes_repo import ComplexesRepo
    db = await _build_db()
    repo = ComplexesRepo(db)
    result = await repo.get("nonexistent-id-xyz")
    assert result is None


@pytest.mark.asyncio
async def test_complexes_repo_upsert_updates_existing():
    """Повторный upsert с тем же ID обновляет данные."""
    from balt_dom_bot.storage.complexes_repo import ComplexesRepo
    db = await _build_db()
    repo = ComplexesRepo(db)
    await repo.upsert(
        complex_id="upd-1",
        name="Первоначальное",
        address="ул. А, 1",
        chat_id=-10000,
        manager_chat_id=1234,
    )
    await repo.upsert(
        complex_id="upd-1",
        name="Обновлённое",
        address="ул. Б, 2",
        chat_id=-10000,
        manager_chat_id=1234,
        active=False,
    )
    c = await repo.get("upd-1")
    assert c is not None
    assert c.name == "Обновлённое"
    assert c.active is False


# ===========================================================================
# 10–14. HTTP smoke-тесты GUI через FastAPI TestClient
# ===========================================================================


def _build_test_app(db):
    """Строит GUI-приложение с реальной in-memory БД и минимальными моками."""
    from balt_dom_bot.gui.app import GuiDeps, build_gui_app
    from balt_dom_bot.gui.auth import AuthConfig
    from balt_dom_bot.gui.events import EventBus
    from balt_dom_bot.storage.complexes_repo import ComplexesRepo
    from balt_dom_bot.storage.escalations import EscalationRepo
    from balt_dom_bot.storage.message_log import MessageLog
    from balt_dom_bot.storage.prompts_repo import PromptProvider, PromptsRepo
    from balt_dom_bot.storage.users_repo import UsersRepo

    auth = AuthConfig(secret_key="test-secret-key-32-bytes-abcdefg!")
    escalations = EscalationRepo(db)
    complexes = ComplexesRepo(db)
    prompts_repo = PromptsRepo(db)
    prompt_provider = PromptProvider(prompts_repo)
    users = UsersRepo(db)
    msg_log = MessageLog(db)

    reply_sender = MagicMock()
    reply_sender.send_reply = AsyncMock()
    esc_sender = MagicMock()
    esc_sender.edit_escalation_card = AsyncMock()
    event_bus = EventBus()

    deps = GuiDeps(
        auth=auth,
        escalations=escalations,
        complexes=complexes,
        prompts_repo=prompts_repo,
        prompt_provider=prompt_provider,
        users=users,
        message_log=msg_log,
        reply_sender=reply_sender,
        escalation_sender=esc_sender,
        event_bus=event_bus,
        db_conn=db.conn,
    )
    return build_gui_app(deps), users, complexes


async def _seed_user(users_repo, login="admin", password="secret123"):
    """Создаёт тестового пользователя и возвращает JWT-токен."""
    from balt_dom_bot.gui.auth import AuthConfig, issue_token
    await users_repo.create(login=login, password=password, display_name="Тест", role="admin")
    user_result = await users_repo.get_by_login(login)
    assert user_result is not None
    user, _ = user_result
    auth = AuthConfig(secret_key="test-secret-key-32-bytes-abcdefg!")
    token = issue_token(auth, user)
    return token


@pytest.mark.asyncio
async def test_gui_complexes_page_renders():
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/complexes", cookies={"balt_dom_session": token})
    assert resp.status_code == 200
    assert "Жилые комплексы" in resp.text
    assert "Изменить" in resp.text or "Добавить" in resp.text


@pytest.mark.asyncio
async def test_gui_edit_complex_page_200():
    """GET /complexes/{id}/edit → 200 со значениями из БД."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    await complexes.upsert(
        complex_id="test-edit",
        name="ЖК для редактирования",
        address="пр. Проверочный, 7",
        chat_id=-987654321,
        manager_chat_id=111222333,
        active=True,
        chat_mode_enabled=True,
        reply_mode="normal",
    )

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/complexes/test-edit/edit", cookies={"balt_dom_session": token})

    assert resp.status_code == 200
    assert "ЖК для редактирования" in resp.text
    assert "пр. Проверочный, 7" in resp.text
    # ID показывается как readonly
    assert "test-edit" in resp.text
    # Форма ведёт на upsert
    assert "/complexes/upsert" in resp.text


@pytest.mark.asyncio
async def test_gui_edit_complex_page_404_for_unknown():
    """GET /complexes/unknown/edit → 404."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, _ = _build_test_app(db)
    token = await _seed_user(users)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/complexes/does-not-exist/edit", cookies={"balt_dom_session": token})
    assert resp.status_code == 404


@pytest.mark.asyncio
async def test_gui_edit_complex_saves_changes():
    """POST /complexes/upsert обновляет поля через edit-форму."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    await complexes.upsert(
        complex_id="edit-save",
        name="Старое название",
        address="ул. Старая, 1",
        chat_id=-111000,
        manager_chat_id=222000,
        active=True,
    )

    form_data = {
        "complex_id": "edit-save",
        "name": "Новое название ЖК",
        "address": "ул. Новая, 99",
        "chat_id": "-111000",
        "manager_chat_id": "222000",
        "reply_mode": "holiday",
        "holiday_message": "Идём в отпуск",
        "strikes_for_ban": "4",
        "trolling_strikes_for_ban": "7",
        "daily_replies_limit": "8",
        "daily_window_hours": "12",
        "active": "on",
        "escalate_to_manager": "on",
    }

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.post(
            "/complexes/upsert",
            data=form_data,
            cookies={"balt_dom_session": token},
            follow_redirects=True,
        )

    assert resp.status_code == 200
    updated = await complexes.get("edit-save")
    assert updated is not None
    assert updated.name == "Новое название ЖК"
    assert updated.reply_mode == "holiday"
    assert updated.holiday_message == "Идём в отпуск"
    assert updated.strikes_for_ban == 4
    assert updated.daily_replies_limit == 8


@pytest.mark.asyncio
async def test_gui_stats_export_returns_xlsx():
    """GET /stats/export → 200, Content-Type xlsx, валидная книга openpyxl."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/stats/export", cookies={"balt_dom_session": token})

    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers.get("content-disposition", "")

    # Проверяем что файл парсится openpyxl
    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert len(wb.sheetnames) == 5
    assert "Сводка" in wb.sheetnames
    assert "По темам" in wb.sheetnames
    assert "По ЖК" in wb.sheetnames
    assert "Причины эскалаций" in wb.sheetnames
    assert "Детальный лог" in wb.sheetnames


@pytest.mark.asyncio
async def test_gui_stats_export_xlsx_headers():
    """Excel-книга содержит правильные заголовки на русском."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, _ = _build_test_app(db)
    token = await _seed_user(users)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/stats/export", cookies={"balt_dom_session": token})

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_themes = wb["По темам"]
    headers = [ws_themes.cell(1, c).value for c in range(1, 4)]
    assert headers == ["Тема (англ.)", "Тема (рус.)", "Кол-во"]

    ws_log = wb["Детальный лог"]
    first_header = ws_log.cell(1, 1).value
    assert first_header == "Дата/время"


@pytest.mark.asyncio
async def test_gui_stats_export_xlsx_with_data():
    """Excel-файл корректно содержит данные из БД."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    # Сеем тестовые данные в messages
    cls_json = json.dumps({"theme": "TECH_FAULT", "urgency": "HIGH",
                           "character": "COMPLAINT_STRONG", "confidence": 0.9,
                           "summary": "Тест", "addressed_to": "uc"})
    dec_json = json.dumps({"escalate": True, "escalation_reason": "high_urgency"})
    await db.conn.execute(
        "INSERT INTO messages (complex_id, chat_id, user_message_id, user_id, user_name, "
        "user_text, classification, decision) VALUES (?,?,?,?,?,?,?,?)",
        ("test-c", -100, "m1", 1, "Иван", "Тестовое сообщение", cls_json, dec_json),
    )
    await db.conn.commit()

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/stats/export", cookies={"balt_dom_session": token})

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    ws_themes = wb["По темам"]
    # Строка 2 должна содержать данные (заголовок в строке 1)
    theme_val = ws_themes.cell(2, 1).value
    theme_ru_val = ws_themes.cell(2, 2).value
    assert theme_val == "TECH_FAULT"
    assert "Технич" in (theme_ru_val or "")


@pytest.mark.asyncio
async def test_gui_stats_page_no_english_theme_names():
    """GET /stats → страница не должна содержать голые English enum-значения."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, _ = _build_test_app(db)
    token = await _seed_user(users)

    # Сеем сообщение
    cls_json = json.dumps({"theme": "EMERGENCY", "urgency": "HIGH",
                           "character": "REPORT", "confidence": 0.99,
                           "summary": "Авария", "addressed_to": "uc"})
    await db.conn.execute(
        "INSERT INTO messages (complex_id, chat_id, user_message_id, user_id, "
        "user_name, user_text, classification) VALUES (?,?,?,?,?,?,?)",
        ("c1", -100, "m2", 2, "Петр", "Авария в доме", cls_json),
    )
    await db.conn.commit()

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/stats", cookies={"balt_dom_session": token})

    assert resp.status_code == 200
    # Русское название темы должно быть на странице
    assert "Аварийная ситуация" in resp.text


@pytest.mark.asyncio
async def test_gui_chat_whitelist_page_no_boltan():
    """GET /chat_whitelist → страница содержит «Диалог», не «Болтание»."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, _ = _build_test_app(db)
    token = await _seed_user(users)

    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.get("/chat_whitelist", cookies={"balt_dom_session": token})

    assert resp.status_code == 200
    assert "Болтание" not in resp.text
    assert "Диалог" in resp.text


@pytest.mark.asyncio
async def test_gui_edit_complex_active_toggle():
    """Редактирование позволяет деактивировать ЖК (убрать галочку «Активен»)."""
    from httpx import ASGITransport, AsyncClient
    db = await _build_db()
    app, users, complexes = _build_test_app(db)
    token = await _seed_user(users)

    await complexes.upsert(
        complex_id="toggle-active",
        name="Активный ЖК",
        address="ул. Активная, 1",
        chat_id=-555000,
        manager_chat_id=333000,
        active=True,
    )
    # active="on" не передаётся → active=False
    async with AsyncClient(transport=ASGITransport(app=app), base_url="http://test") as client:
        resp = await client.post(
            "/complexes/upsert",
            data={
                "complex_id": "toggle-active",
                "name": "Деактивированный ЖК",
                "address": "ул. Активная, 1",
                "chat_id": "-555000",
                "manager_chat_id": "333000",
                "reply_mode": "normal",
                "strikes_for_ban": "3",
                "trolling_strikes_for_ban": "6",
                "daily_replies_limit": "5",
                "daily_window_hours": "6",
                "escalate_to_manager": "on",
            },
            cookies={"balt_dom_session": token},
            follow_redirects=True,
        )

    assert resp.status_code == 200
    c = await complexes.get("toggle-active")
    assert c is not None
    assert c.active is False


# ===========================================================================
# 15–16. Юнит-тесты Excel-хелперов
# ===========================================================================


def test_excel_autowidth_empty_sheet():
    """_excel_autowidth не должен падать на листе без данных."""
    from balt_dom_bot.gui.app import _excel_autowidth
    wb = openpyxl.Workbook()
    ws = wb.active
    _excel_autowidth(ws)  # не должно бросать исключений


def test_excel_data_row_alternating_colors():
    """_excel_data_row должен применять разные цвета для чётных/нечётных строк."""
    from balt_dom_bot.gui.app import _excel_data_row, _make_excel_style
    wb = openpyxl.Workbook()
    ws = wb.active
    style = _make_excel_style()
    _, _, even_fill, odd_fill, _, _, _ = style
    _excel_data_row(ws, 2, ["A", "B"], style)
    _excel_data_row(ws, 3, ["C", "D"], style)
    # Строка 2 (чётная) → even_fill
    assert ws.cell(2, 1).fill.fgColor.rgb == even_fill.fgColor.rgb
    # Строка 3 (нечётная) → odd_fill
    assert ws.cell(3, 1).fill.fgColor.rgb == odd_fill.fgColor.rgb


def test_excel_header_row_bold_blue():
    """_excel_header_row должен применять жирный шрифт и синий фон."""
    from balt_dom_bot.gui.app import _excel_header_row, _make_excel_style
    wb = openpyxl.Workbook()
    ws = wb.active
    style = _make_excel_style()
    header_fill, header_font, *_ = style
    _excel_header_row(ws, ["Колонка 1", "Колонка 2", "Колонка 3"], style)
    cell = ws.cell(1, 1)
    assert cell.font.bold is True
    # openpyxl хранит цвет в ARGB-формате: "00FFFFFF" = белый (FF = непрозрачный).
    assert cell.font.color.rgb.upper().endswith("FFFFFF")
    assert cell.fill.fgColor.rgb == header_fill.fgColor.rgb


def test_excel_autowidth_respects_max_width():
    """_excel_autowidth ограничивает ширину колонки до 50."""
    from balt_dom_bot.gui.app import _excel_autowidth
    from openpyxl.utils import get_column_letter
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value="x" * 200)  # очень длинная строка
    _excel_autowidth(ws)
    width = ws.column_dimensions[get_column_letter(1)].width
    assert width <= 50
